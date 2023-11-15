import json
from django.shortcuts import render
from .models import VetUsers, VetClient
from django.http import HttpResponse
from django.contrib import messages

from openpyxl import Workbook
from django.http import HttpResponse
# Create your views here.
def stopVet(request, pk):

    user = VetUsers.objects.get(id=pk)
    status = user.status
    user.status = not status
    user.save()
    return HttpResponse(json.dumps(
            {
                "status":str(status),
            }
        ),
        content_type="application/json"
    )

def clearVet(request):

    messages.success(request, 'veterenaliya tarixi tozalandi')

    VetClient.objects.all().delete()
    return HttpResponse(json.dumps(
            {
                "status":True,
            }
        ),
        content_type="application/json"
    )


def export_excel(request):

    vusers = VetUsers.objects.all()
    for i in vusers:
        # Create a workbook and add a worksheet to it
        wb = Workbook()
        ws = wb.active

        # Write the headers to the worksheet
        headers = ["Xizmat ko'rsadi", 'Name', 'Phone', 'Location', 'Day', 'Humidity', 'Temperature', 'Sickness', 'Diagnose', 'Img']  # Replace with your field names

        for col_num, column_title in enumerate(headers, 1):
            col_letter = ws.cell(row=1, column=col_num).column_letter
            ws['{}1'.format(col_letter)] = column_title
            ws.column_dimensions[col_letter].width = 15

        # Write the data to the worksheet
        for row_num, obj in enumerate(VetClient.objects.all(), 2):
            ws.cell(row=row_num, column=1, value=i.name)  # Replace 'field1' with your field name
            ws.cell(row=row_num, column=1, value=obj.name)  # Replace 'field1' with your field name
            ws.cell(row=row_num, column=2, value=obj.phone)  # Replace 'field2' with your field name
            ws.cell(row=row_num, column=3, value=obj.location)  # Replace 'field3' with your field name
            ws.cell(row=row_num, column=4, value=obj.day)  # Replace 'field3' with your field name
            ws.cell(row=row_num, column=5, value=obj.humidity)  # Replace 'field3' with your field name
            ws.cell(row=row_num, column=6, value=obj.temperature)  # Replace 'field3' with your field name
            ws.cell(row=row_num, column=7, value=obj.sickness)  # Replace 'field3' with your field name
            ws.cell(row=row_num, column=8, value=obj.diagnose)  # Replace 'field3' with your field name
            ws.cell(row=row_num, column=9, value=obj.img.url)  # Replace 'field3' with your field name
            ws.cell(row=row_num, column=9, value=obj.create_at)  # Replace 'field3' with your field name

    # Create the HttpResponse object
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=your_file_name.xlsx'
    wb.save(response)

    return response