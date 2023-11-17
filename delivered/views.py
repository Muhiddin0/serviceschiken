import json
from pyexpat.errors import messages
from django.http import HttpResponse
from django.shortcuts import render
from .models import DeliverUsers, Delivered

from openpyxl import Workbook
from django.http import HttpResponse
# Create your views here.

# Create your views here.
def stopDelivred(request, pk):

    user = DeliverUsers.objects.get(id=pk)
    status = user.status
    user.status = not status
    user.save()
    return HttpResponse(json.dumps(
            {
                "status":str(status),
            }
        ),
        content_type="application/json"
    )


def clear(request):

    messages.success(request, 'veterenaliya tarixi tozalandi')

    Delivered.objects.all().delete()
    return render(request, 'delivred.html')


def export_excel(request):
    # Create a workbook and add a worksheet to it
    wb = Workbook()
    ws = wb.active

    duser = DeliverUsers.objects.all()
    for i in duser:

        # Write the headers to the worksheet
        headers = ["Xizmat ko'rsatdi", 'Name', 'Phone', 'Location', 'price', 'img', 'comment_img', 'comment']  # Replace with your field names
        for col_num, column_title in enumerate(headers, 1):
            col_letter = ws.cell(row=1, column=col_num).column_letter
            ws['{}1'.format(col_letter)] = column_title
            ws.column_dimensions[col_letter].width = 15

        # Write the data to the worksheet
        for row_num, obj in enumerate(Delivered.objects.all(), 2):
            ws.cell(row=row_num, column=1, value=i.name)  # Replace 'field1' with your field name
            ws.cell(row=row_num, column=1, value=obj.name)  # Replace 'field1' with your field name
            ws.cell(row=row_num, column=2, value=obj.phone)  # Replace 'field2' with your field name
            ws.cell(row=row_num, column=3, value=obj.location)  # Replace 'field3' with your field name
            ws.cell(row=row_num, column=4, value=obj.price)  # Replace 'field3' with your field name
            ws.cell(row=row_num, column=5, value=obj.img)  # Replace 'field3' with your field name
            ws.cell(row=row_num, column=6, value=obj.comment_img)  # Replace 'field3' with your field name
            ws.cell(row=row_num, column=7, value=obj.comment)  # Replace 'field3' with your field name
            ws.cell(row=row_num, column=7, value=obj.create_at)  # Replace 'field3' with your field name

    # Create the HttpResponse object
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=your_file_name.xlsx'
    wb.save(response)

    return response